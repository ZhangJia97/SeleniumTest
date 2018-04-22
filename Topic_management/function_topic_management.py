from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from time import sleep
import datetime
from openpyxl import Workbook,load_workbook

#新建一个报告
def create_xlsx():
    # 打开名字为Topic_management+测试时间的表格
    wb = Workbook()
    ws = wb.active
    xlsx_name = "Topic_management.xlsx"
    test_about = "开始进行题库管理的自动化测试，本次自动化测试时间为" + str(datetime.datetime.now())
    ws["A1"] = test_about
    ws["A2"] = "页面对应的url"
    ws["B2"] = "输入数据/操作方式"
    ws["C2"] = "预计出现结果"
    ws["D2"] = "实际出现结果"
    wb.save(xlsx_name)

def read_choice_xlsx(choice_num):
    wb_read = load_workbook("./input_data/choice_list_data.xlsx")
    ws_read = wb_read.active
    stem = ws_read[str("A"+str(choice_num))].value
    choice_a = ws_read[str("B"+str(choice_num))].value
    choice_b = ws_read[str("C"+str(choice_num))].value
    choice_c = ws_read[str("D"+str(choice_num))].value
    choice_d = ws_read[str("E"+str(choice_num))].value
    right_choice = ws_read[str("F"+str(choice_num))].value
    course = ws_read[str("G"+str(choice_num))].value
    knowledge_one = ws_read[str("H"+str(choice_num))].value
    knowledge_two = ws_read[str("I"+str(choice_num))].value
    expected_results = ws_read[str("J"+str(choice_num))].value
    row = ws_read.max_row
    input_data = {
        'stem':stem,
        'choice_a':choice_a,
        'choice_b':choice_b,
        'choice_c':choice_c,
        'choice_d':choice_d,
        'right_choice':right_choice,
        'course':course,
        'knowledge_one':knowledge_one,
        'knowledge_two':knowledge_two,
        'expected_results':expected_results,
        'row':row
    }
    return input_data

def read_problem_xlsx(choice_num,choice_list):
    #本函数用来读取非选择题的题目的数据
    xlsx_file_name = './input_data/'+ choice_list + '_data.xlsx'
    wb_read = load_workbook(xlsx_file_name)
    ws_read = wb_read.active
    title = ws_read[str("A"+str(choice_num))].value
    course = ws_read[str("B"+str(choice_num))].value
    knowledge_one = ws_read[str("C"+str(choice_num))].value
    knowledge_two = ws_read[str("D"+str(choice_num))].value
    topic_description = ws_read[str("E"+str(choice_num))].value
    input_description = ws_read[str("F"+str(choice_num))].value
    output_description = ws_read[str("G"+str(choice_num))].value
    input_example_1 = ws_read[str("H"+str(choice_num))].value
    input_example_2 = ws_read[str("I"+str(choice_num))].value
    output_example_1 = ws_read[str("J"+str(choice_num))].value
    output_example_2 = ws_read[str("K"+str(choice_num))].value
    if ws_read[str("L"+str(choice_num))].value == "是":
        test_file_name = "sample.zip"
    else:
        test_file_name = "error.zip"
    expected_results = ws_read[str("M"+str(choice_num))].value
    row = ws_read.max_row
    if choice_list == "biancheng_list":
        input_data = {
            'title':title,
            'course':course,
            'knowledge_one':knowledge_one,
            'knowledge_two':knowledge_two,
            'topic_description':topic_description,
            'input_description':input_description,
            'output_description':output_description,
            'input_example_1':input_example_1,
            'input_example_2':input_example_2,
            'output_example_1':output_example_1,
            'output_example_2':output_example_2,
            'test_file_name':test_file_name,
            'expected_results':expected_results,
            'row':row
        }
    else:
        code = ws_read[str("N"+str(choice_num))].value
        input_data = {
            'title': title,
            'course': course,
            'knowledge_one': knowledge_one,
            'knowledge_two': knowledge_two,
            'topic_description': topic_description,
            'input_description': input_description,
            'output_description': output_description,
            'input_example_1': input_example_1,
            'input_example_2': input_example_2,
            'output_example_1': output_example_1,
            'output_example_2': output_example_2,
            'test_file_name': test_file_name,
            'expected_results': expected_results,
            'row': row,
            'code': code
        }
    return input_data



def add_xlsx(text_now,text_type,i):
    wb = load_workbook("Topic_management.xlsx")
    ws = wb.active
    ws1 = wb.get_sheet_by_name("Sheet")
    num = str(text_type + str(i))
    ws1[num] = text_now
    wb.save("Topic_management.xlsx")

# 添加cookie
def add_cookies(driver):
    # 更改浏览器cookie做到免密登录
    cookies = [
        {'name': 'sessionid', 'value': '5gta4vuy9uncnqdxoh9wlbxs8yqje2kr'},
        {'name': 'csrftoken', 'value': 'zYuD5vaObDF8lFeUPlgyj6eONtzZjTW2'},
    ]
    driver.delete_all_cookies()
    for cookie in cookies:
        driver.add_cookie(cookie)


class Problem_list_test():
    # 类中含有的是测试题库管理功能的各个函数
    def __init__(self,driver,choice_list,test_num,input_data):
        self.driver = driver
        self.choice_list = choice_list
        self.test_num = test_num
        self.input_data = input_data


    def is_element_exist(self,css):
        # 本函数主要判断页面是否含有某个元素
        s = self.driver.find_element_by_css_selector(css_selector=css).text
        if s:
            return True
        else:
            return False


    def test_select_knowledge(self):
        # 本函数主要是点击测试各个知识点

        #将实验报告保存于文件中
        url_now = self.driver.current_url
        add_xlsx(url_now, "A", self.test_num)
        test_now = "将"+self.choice_list+"中的知识点进行遍历"
        add_xlsx(test_now, "B", self.test_num)
        test_now = "成功遍历知识点"
        add_xlsx(test_now, "C", self.test_num)

        select1 = Select(self.driver.find_element_by_css_selector('#classname'))
        js_select1 = "return document.getElementById('classname').options.length"  # 获取select中options的个数
        options = self.driver.execute_script(js_select1)
        if self.driver.find_element_by_css_selector('#table > tbody > tr > td').text == "没有找到匹配的记录":
            data_error = True
        else:
            data_error = False
        for x1 in range(1, options):
            select1.select_by_index(x1)
            sleep(1)
            select2 = Select(self.driver.find_element_by_css_selector('#id_knowledgePoint1'))
            js_select2 = "return document.getElementById('id_knowledgePoint1').options.length"
            options2 = self.driver.execute_script(js_select2)
            for x2 in range(1, options2):
                select2.select_by_index(x2)
                sleep(1)
                select3 = Select(self.driver.find_element_by_css_selector('#id_knowledgePoint2'))
                js_select3 = "return document.getElementById('id_knowledgePoint2').options.length"
                options3 = self.driver.execute_script(js_select3)
                for x3 in range(1, options3):
                    select3.select_by_index(x3)
                    self.driver.find_element_by_css_selector('#ok').click()
                    sleep(1)
                    data_exist = self.driver.find_element_by_xpath('//*[@id="table"]/tbody/tr/td').text
                    if(data_exist=="没有找到匹配的记录" and data_error==False):
                        data_error = False
                    else:
                        data_error = True
                    sleep(1)

        # 将实验报告保存于文件中
        if data_error == True:
            test_now = "成功遍历知识点"
        else:
            test_now = "未能成功遍历知识点"
        add_xlsx(test_now, "D", self.test_num)
        self.test_num += 1


    def test_add(self):
        # 本函数主要测试题库的添加功能

        url_now = self.driver.current_url
        add_xlsx(url_now, "A", self.test_num)
        test_now = "输入的数据为：" + str(self.input_data)
        add_xlsx(test_now, "B", self.test_num)
        test_now = self.input_data['expected_results']
        add_xlsx(test_now, "C", self.test_num)

        if self.choice_list != 'choice_list':           #由于填空题与其他的三个题型格式不同，所以需要在此处判断题型
            # 输入内容
            try:
                self.driver.find_element_by_css_selector('#id_title').send_keys(self.input_data['title'])  # 标题
            except TypeError:
                pass

            # 选择知识点
            try:
                boancheng1 = Select(self.driver.find_element_by_css_selector('#id_classname'))
                boancheng1.select_by_index(int(self.input_data['course']))  # 1指具体的选项
            except TypeError:
                pass
            sleep(1)

            try:
                boancheng2 = Select(self.driver.find_element_by_css_selector('#id_knowledgePoint1'))
                boancheng2.select_by_index(int(self.input_data['knowledge_one']))
            except TypeError:
                pass
            sleep(1)

            try:
                boancheng3 = Select(self.driver.find_element_by_css_selector('#id_knowledgePoint2'))
                boancheng3.select_by_index(int(self.input_data['knowledge_two']))
            except TypeError:
                pass
            sleep(1)

            try:
                self.driver.find_element_by_css_selector('#id_description').send_keys(
                    self.input_data['topic_description'])  # 题目描述
            except TypeError:
                pass

            try:
                self.driver.find_element_by_css_selector(
                    '#problem-form > div:nth-child(7) > div:nth-child(4) > button').click()  # 放在这里点击添加是因为当点好知识点二之后直接点添加会有报错
            except TypeError:
                pass

            try:
                self.driver.find_element_by_css_selector('#id_input').send_keys(self.input_data['input_description'])
            except TypeError:
                pass

            try:
                self.driver.find_element_by_css_selector('#id_output').send_keys(self.input_data['output_description'])
            except TypeError:
                pass

            if self.choice_list != 'biancheng_list':
                try:
                    self.driver.find_element_by_css_selector('#id_sample_code').send_keys(
                        self.input_data['code'])
                except TypeError:
                    pass

            try:
                self.driver.find_element_by_css_selector('#id_sample_input1').send_keys(
                    self.input_data['input_example_1'])
            except TypeError:
                pass

            try:
                self.driver.find_element_by_css_selector('#id_sample_output1').send_keys(
                    self.input_data['output_example_1'])
            except TypeError:
                pass

            try:
                self.driver.find_element_by_css_selector('#id_sample_input2').send_keys(
                    self.input_data['input_example_2'])
            except TypeError:
                pass

            try:
                self.driver.find_element_by_css_selector('#id_sample_output2').send_keys(
                    self.input_data['output_example_2'])
            except TypeError:
                pass

            self.driver.find_element_by_xpath('//*[@id="problem-form"]/button[1]').click()  # 上传测试用例
            test_file_url = '/Users/zhangjia/PycharmProjects/test_problem/Topic_management/' + self.input_data['test_file_name']#当实际运行是请将本路径改为实际的测试用例文件的路径
            self.driver.find_element_by_css_selector('#input-43').send_keys(test_file_url)  # 上传文件
            sleep(1)

            self.driver.find_element_by_css_selector(
                '#myModal > div > div > div.modal-body > div.file-input > div.input-group.file-caption-main > div.input-group-btn > a').click()  # 点击上传验证
            sleep(1)

            # 因为在测试编程题时如果测试用例错误保存时将会有一个弹窗警告，所以此处判断题型
            if self.driver.find_element_by_css_selector('#alert-info').text != '文件有效！':
                self.driver.find_element_by_css_selector('#close').click()
                sleep(1)
                self.driver.find_element_by_css_selector('#problem-form > button.btn.btn-success.btn-lg').click()
                sleep(1)
                self.driver.switch_to_alert().accept()
            else:
                self.driver.find_element_by_css_selector('#close').click()
                sleep(1)
                self.driver.find_element_by_css_selector('#problem-form > button.btn.btn-success.btn-lg').click()

            if self.choice_list == 'biancheng_list':
                if self.driver.find_element_by_xpath('//*[@id="main-content"]/section/h3').text == '编程题“'+str(self.input_data['title'])+'”的详细信息':
                    test_now = self.input_data['expected_results']
                elif self.input_data['expected_results'] == '题目添加失败':
                    test_now = self.input_data['expected_results']
                else:
                    test_now = "未能" + self.input_data['expected_results']
                add_xlsx(test_now, "D", self.test_num)
            elif self.choice_list == 'tiankong_list':
                if self.driver.find_element_by_xpath('//*[@id="main-content"]/section/h3').text == '程序填空题“'+str(self.input_data['title'])+'”的详细信息':
                    test_now = self.input_data['expected_results']
                elif self.input_data['expected_results'] == '题目添加失败':
                    test_now = self.input_data['expected_results']
                else:
                    test_now = "未能" + self.input_data['expected_results']
                add_xlsx(test_now, "D", self.test_num)
            else:
                if self.driver.find_element_by_xpath('//*[@id="main-content"]/section/h3').text == '程序改错题“'+str(self.input_data['title'])+'”的详细信息':
                    test_now = self.input_data['expected_results']
                elif self.input_data['expected_results'] == '题目添加失败':
                    test_now = self.input_data['expected_results']
                else:
                    test_now = "未能" + self.input_data['expected_results']
                add_xlsx(test_now, "D", self.test_num)
            self.test_num += 1

        else:
            # 开始向选择题中输入具体的内容
            sleep(1)
            try:
                self.driver.find_element_by_css_selector('#id_title').send_keys(self.input_data['stem'])  # 输入题干
            except TypeError:
                pass

            # 输入四个选项
            try:
                self.driver.find_element_by_css_selector('#id_a').send_keys(self.input_data['choice_a'])
            except TypeError:
                pass

            try:
                self.driver.find_element_by_css_selector('#id_b').send_keys(self.input_data['choice_b'])
            except TypeError:
                pass

            try:
                self.driver.find_element_by_css_selector('#id_c').send_keys(self.input_data['choice_c'])
            except TypeError:
                pass

            try:
                self.driver.find_element_by_css_selector('#id_d').send_keys(self.input_data['choice_d'])
            except TypeError:
                pass


            # 选择正确选项
            try:
                right_choice = "#id_selection_" + str(self.input_data['right_choice'] - 1)
                self.driver.find_element_by_css_selector(right_choice).click()
            except TypeError:
                pass
            sleep(1)

            try:
                xuanze_select1 = Select(self.driver.find_element_by_css_selector('#id_classname'))
                xuanze_select1.select_by_index(int(self.input_data['course']))  # 1指具体的选项
            except TypeError:
                pass
            sleep(1)

            try:
                xuanze_select2 = Select(self.driver.find_element_by_css_selector('#id_knowledgePoint1'))
                xuanze_select2.select_by_index(int(self.input_data['knowledge_one']))
            except TypeError:
                pass
            sleep(1)

            try:
                xuanze_select3 = Select(self.driver.find_element_by_css_selector('#id_knowledgePoint2'))
                xuanze_select3.select_by_index(int(self.input_data['knowledge_two']))
                self.driver.find_element_by_css_selector(
                    '#main-content > section > div > div > div > div > div > form > div.form-group.row > div:nth-child(4) > button').click()
            except TypeError:
                pass
            sleep(1)

            # 获取各个空的内容
            tigan = self.driver.find_element_by_css_selector('#id_title').text
            xuanxiang = self.driver.find_element_by_css_selector('#id_selection_0').is_selected() or \
                        self.driver.find_element_by_css_selector('#id_selection_1').is_selected() or \
                        self.driver.find_element_by_css_selector('#id_selection_2').is_selected() or \
                        self.driver.find_element_by_css_selector('#id_selection_3').is_selected()
            a_choice = self.driver.find_element_by_css_selector('#id_a').text
            b_choice = self.driver.find_element_by_css_selector('#id_b').text
            c_choice = self.driver.find_element_by_css_selector('#id_c').text
            d_choice = self.driver.find_element_by_css_selector('#id_d').text

            # 判断有没有未输入或者选择的空，由于type验证原因，未输入的空将自动输入一个空格
            if tigan == '':
                self.driver.find_element_by_css_selector('#id_title').send_keys(' ')
            if xuanxiang == False:
                self.driver.find_element_by_css_selector('#id_selection_0').click()  # 若未选择正确选项，则默认为A
            if a_choice == '':
                self.driver.find_element_by_css_selector('#id_a').send_keys(' ')
            if b_choice == '':
                self.driver.find_element_by_css_selector('#id_b').send_keys(' ')
            if c_choice == '':
                self.driver.find_element_by_css_selector('#id_c').send_keys(' ')
            if d_choice == '':
                self.driver.find_element_by_css_selector('#id_d').send_keys(' ')

            self.driver.find_element_by_xpath('//*[@id="main-content"]/section/div/div/div/div/div/form/p/button').click()
            sleep(1)

            if self.driver.find_element_by_xpath('//*[@id="main-content"]/section/h3').text == '选择题“'+str(self.input_data['stem'])+'”的详细信息':
                test_now = self.input_data['expected_results']
            elif self.input_data['expected_results']=="新建选择题失败":
                test_now = self.input_data['expected_results']
            else:
                test_now = "未能"+self.input_data['expected_results']
            add_xlsx(test_now, "D", self.test_num)
            self.test_num += 1


    def test_pages(self):
        # 测试下一页按钮
        # 这个是通过判断每页显示多少条的按钮是否存在来判断是否点击下一页

        url_now = self.driver.current_url
        add_xlsx(url_now, "A", self.test_num)
        test_now = "测试下一页按钮是否可用"
        add_xlsx(test_now, "B", self.test_num)
        test_now = "下一页按钮可用"
        add_xlsx(test_now, "C", self.test_num)
        sleep(1)

        #由于当题目较少时判断不能根据css_selection判断，只能通过循环，找到最大页码
        if self.is_element_exist('#main-content > section > div > div > div > div.bootstrap-table > div.fixed-table-container > div.fixed-table-pagination > div.pull-left.pagination-detail > span.page-list > span > button > span.page-size'):
            sleep(2)
            li_num = 6
            pages = None
            try:
                pages = self.driver.find_element_by_css_selector(
                    '#main-content > section > div > div > div > div.bootstrap-table > div.fixed-table-container > div.fixed-table-pagination > div.pull-right.pagination > ul > li.page-last > a').text
            except:
                while True:
                    li_xpath_selector = '// *[ @ id = "main-content"] / section / div / div / div / div[2] / div[2] / div[4] / div[2] / ul / li['+str(li_num)+'] / a'

                    try:
                        pages = self.driver.find_element_by_xpath(li_xpath_selector).text
                    except:
                        pass

                    if pages != None and pages != '›':
                        break
                    else:
                        li_num -= 1

            try:
                for x in range(0, int(pages)):
                    self.driver.find_element_by_css_selector(
                        '#main-content > section > div > div > div > div.bootstrap-table > div.fixed-table-container > div.fixed-table-pagination > div.pull-right.pagination > ul > li.page-next > a').click()
                    sleep(1)
                test_now = "下一页按钮可用"
                add_xlsx(test_now, "D", self.test_num)
            except:
                test_now = "下一页按钮不可用"
                add_xlsx(test_now, "D", self.test_num)