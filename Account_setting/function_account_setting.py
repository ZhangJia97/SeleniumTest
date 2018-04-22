from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from time import sleep
import datetime
from openpyxl import Workbook,load_workbook

#新建一个报告
def create_xlsx():
    # 打开名字为Account_setting的表格
    wb = Workbook()
    ws = wb.active
    xlsx_name = "Account_setting.xlsx"
    test_about = "开始进行账号相关的自动化测试，本次自动化测试时间为" + str(datetime.datetime.now())
    ws["A1"] = test_about
    ws["A2"] = "页面对应的url"
    ws["B2"] = "输入数据/操作方式"
    ws["C2"] = "预计出现结果"
    ws["D2"] = "实际出现结果"
    wb.save(xlsx_name)

def read_register_xlsx(register_num):
    wb_read = load_workbook("./input_data/register_data.xlsx")
    ws_read = wb_read.active
    email = ws_read[str("A"+str(register_num))].value
    nick_name = ws_read[str("B"+str(register_num))].value
    user_name = ws_read[str("C"+str(register_num))].value
    password = ws_read[str("D"+str(register_num))].value
    confirmation_password = ws_read[str("E"+str(register_num))].value
    expected_results = ws_read[str("F"+str(register_num))].value
    row = ws_read.max_row
    input_data = {
        'email':email,
        'nick_name':nick_name,
        'user_name':user_name,
        'password':password,
        'confirmation_password':confirmation_password,
        'expected_results':expected_results,
        'row':row
    }
    return input_data

def read_login_xlsx(register_num):
    wb_read = load_workbook("./input_data/login_data.xlsx")
    ws_read = wb_read.active
    user_name = ws_read[str("A"+str(register_num))].value
    password = ws_read[str("B"+str(register_num))].value
    expected_results = ws_read[str("C"+str(register_num))].value
    row = ws_read.max_row
    input_data = {
        'user_name':user_name,
        'password':password,
        'expected_results':expected_results,
        'row':row
    }
    return input_data

def add_xlsx(text_now,text_type,i):
    wb = load_workbook("Account_setting.xlsx")
    ws = wb.active
    ws1 = wb.get_sheet_by_name("Sheet")
    num = str(text_type + str(i))
    ws1[num] = text_now
    wb.save("Account_setting.xlsx")

# 添加cookie
def add_cookies(driver):
    # 更改浏览器cookie做到免密登录
    cookies = [
        {'name': 'sessionid', 'value': 'nl013yfz073sn54r0vvkygv9687gidbl'},
        {'name': 'csrftoken', 'value': 'M1DPybCQQljZVwRrHNvCsowA1Uws5jTL'},
    ]
    driver.delete_all_cookies()
    for cookie in cookies:
        driver.add_cookie(cookie)


class Register_test():
    # 类中含有测试注册功能的函数
    def __init__(self,driver,test_num,input_data):
        self.driver = driver
        self.test_num = test_num
        self.input_data = input_data

    def test_register(self):
        # 本函数主要用来测试注册
        url_now = self.driver.current_url
        add_xlsx(url_now, "A", self.test_num)
        test_now = "输入的数据为：" + str(self.input_data)
        add_xlsx(test_now,"B",self.test_num)
        test_now = self.input_data['expected_results']
        add_xlsx(test_now, "C", self.test_num)
        js = "var elem = document.getElementById('vmaig-auth-register-email');elem.getAttribute('type');elem.setAttribute('type','')"
        self.driver.execute_script(js)
        self.driver.find_element_by_id('vmaig-auth-register-email').send_keys(self.input_data['email'])
        self.driver.find_element_by_id('vmaig-auth-register-username').send_keys(self.input_data['nick_name'])
        self.driver.find_element_by_id('vmaig-auth-register-id').send_keys(self.input_data['user_name'])
        self.driver.find_element_by_id('vmaig-auth-register-password1').send_keys(self.input_data['password'])
        self.driver.find_element_by_id('vmaig-auth-register-password2').send_keys(self.input_data['confirmation_password'])
        self.driver.find_element_by_id('vmaig-auth-register-button').click()
        sleep(1)
        if self.driver.title == '注册' and self.input_data['expected_results'] == '注册失败':
            test_now = self.input_data['expected_results']
        elif self.driver.title == '注册' and self.input_data['expected_results'] != '注册失败':
            test_now = "未能" + str(self.input_data['expected_results'])
        else:
            test_now = self.input_data['expected_results']
        add_xlsx(test_now, "D", self.test_num)



class Login_test():
    # 类中含有测试登录功能的函数
    def __init__(self,driver,test_num,input_data):
        self.driver = driver
        self.test_num = test_num
        self.input_data = input_data

    def test_login(self):
        # 本函数用来测试登录功能
        # 输入用户名和密码
        url_now = self.driver.current_url
        add_xlsx(url_now, "A", self.test_num)
        test_now = "输入的数据为：" + str(self.input_data)
        add_xlsx(test_now, "B", self.test_num)
        test_now = self.input_data['expected_results']
        add_xlsx(test_now, "C", self.test_num)
        self.driver.find_element_by_id('vmaig-auth-login-username').send_keys(self.input_data['user_name'])
        self.driver.find_element_by_id("vmaig-auth-login-password").send_keys(self.input_data['password'])
        self.driver.find_element_by_id("vmaig-auth-login-button").click()
        sleep(1)
        # 用户名和密码如果错误，输出错误，反之，输出登录成功
        '''
            由于经过测试发现此处用if，else时当输入账户密码错误时会出现弹窗此时将无法使用driver.title()
            但是当使用try时将没有这个情况，为了代码简洁所以用try
        '''
        try:
            if self.driver.title != '登录' and self.input_data['expected_results'] == '登录成功':
                test_now = self.input_data['expected_results']
                self.driver.find_element_by_css_selector(
                    '#nav-accordion > li:nth-child(9) > a > span:nth-child(2)').click()
                sleep(2)
                self.driver.find_element_by_css_selector(
                    '#nav-accordion > li:nth-child(9) > ul > li:nth-child(2) > a').click()
                sleep(2)
        except:
            alter = self.driver.switch_to_alert()
            try:
                if self.driver.title == '登录' and self.input_data['expected_results'] == '登录失败':
                    test_now = self.input_data['expected_results']
                else:
                    test_now = '未能' + self.input_data['expected_results']
            except:
                alter.accept()
        add_xlsx(test_now,'D',self.test_num)

class Cancellation_account():
    # 类中含有测试注销功能的函数
    def __init__(self,driver,test_num):
        self.driver = driver
        self.test_num = test_num

    def test_cancellation_account(self):
        url_now = self.driver.current_url
        add_xlsx(url_now, "A", self.test_num)
        test_now = "点击注销按钮"
        add_xlsx(test_now, "B", self.test_num)
        test_now = "注销成功"
        add_xlsx(test_now, "C", self.test_num)
        self.driver.find_element_by_css_selector('#nav-accordion > li:nth-child(9) > a > span:nth-child(2)').click()
        sleep(2)
        self.driver.find_element_by_css_selector('#nav-accordion > li:nth-child(9) > ul > li:nth-child(2) > a').click()
        sleep(2)
        if self.driver.find_element_by_css_selector('#top_menu > ul > li:nth-child(2) > a').text == '登录':
            test_now = "注销成功"
        else:
            test_now = "未能注销成功"
        add_xlsx(test_now,"D",self.test_num)